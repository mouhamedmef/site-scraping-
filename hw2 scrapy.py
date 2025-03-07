#MOHAMED MEFTAH TD3TP2
import requests
from bs4 import BeautifulSoup
import os
from openpyxl import Workbook, load_workbook
script_dir = os.path.dirname(os.path.abspath(__file__))  
excel_filename = os.path.join(script_dir, 'match_resultat.xlsx')  # Save file in the same directory

journee = input("Entrez numero de la journee: ")

url = f"https://www.ftf.org.tn/fr/calendrier-et-resultats-ligue-1/?journee={journee}&season=2024-2025"
page = requests.get(url)

def scrape_page(page):
    src = page.content
    soup = BeautifulSoup(src, 'lxml')
    journee_title = soup.find("div", class_="journee-title")
    num_journee = journee_title.text.strip() if journee_title else f"Journee {journee}"
    print(f"Journee: {num_journee}")
    
    
    match_containers = soup.find_all("div", class_="leTous")  
    
    if not match_containers:
        print("No matches found for this journee.")
        return
    
    
    file_exists = os.path.exists(excel_filename)
    
    if file_exists:
        wb = load_workbook(excel_filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Journee", "Match", "Resultat"])  
    
    for match in match_containers:
        teams = match.find_all("div", class_="match_eq_title")  
        if len(teams) < 2:
            continue  
        
        result_divs = match.find_all("div", class_="match_res")  
        
        for i in range(0, len(teams) - 1, 2):  # Getting teams 2 by 2
            team_a = teams[i].text.strip()
            team_b = teams[i + 1].text.strip()
            match_name = f"{team_a} vs {team_b}"
            
            result = result_divs[i // 2].text.strip() if i // 2 < len(result_divs) else "N/A"
            
            print(f"Adding: {match_name} - Result: {result}")
            
            ws.append([num_journee, match_name, result])  # Append to Excel file
    wb.save(excel_filename)
    print(f"All matches for {num_journee} saved to {excel_filename}")

scrape_page(page)